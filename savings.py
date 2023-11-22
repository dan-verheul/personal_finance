#import setup
from folder_constants import *

#pick the needed df's from folder_constants.py
method = 'savings'
original_output_data = original_output_dataframes[method]
upload_df = upload_df_dictionary[method]

#update values
def update_description(description):
    if any(substring in description for substring in ['Requested transfer from ALLY BANK Spending account XXXXXX8706', 
                                                      'Internet transfer from Spending account XXXXXX8706']):
        return 'Transfer from Ally Checking'
    if any(substring in description for substring in ['Requested transfer to ALLY BANK Spending account XXXXXX8706', 
                                                      'Internet transfer to Spending account XXXXXX8706']):
        return 'Transfer to Ally Checking'
    elif 'EOS FITNESS' in description:
        return 'EOS Gym'
    elif any(substring in description for substring in ['ATM Fee Reimbursement', 'Interest Paid']):
        return description
    elif description == 'NFCU ACH PAYMENT':
        return 'NFCU Credit Card Payment'
    else:
        return '**'+description+'**'

def update_df(df):
    for index, row in df.iterrows():
        df.at[index, 'Description'] = update_description(row['Description'])

update_df(upload_df)

# format, add combo col
original_output_data, upload_df = format_and_combo(original_output_data, upload_df)

# remove rows already stored in output sheet so they're not uploaded twice
upload_df = remove_rows_already_saved(original_output_data,upload_df)

# now join the upload_df with the original_df to include original rows
upload_df = pd.concat([upload_df,original_output_data],ignore_index=True)
if 'combo' in upload_df.columns:
            upload_df = upload_df.drop('combo', axis=1)

#append leftover rows to google sheets
if len(upload_df) > 0:
    #format and sort
    upload_df['Date'] = pd.to_datetime(upload_df['Date'])
    upload_df = upload_df.sort_values(by=['Date']).reset_index(drop=True)
    upload_df['Date'] = upload_df['Date'].astype(str)

    # Update Google Sheets
    # clear
    worksheet = spreadsheet.worksheet('Outputs')
    range_to_clear = worksheet.range(sheets_info_dict[method][1])
    for cell in range_to_clear:
        cell.value = ''
    worksheet.update_cells(range_to_clear)
    #upload
    data = upload_df.values.tolist()
    worksheet.update(sheets_info_dict[method][0], data)


    # Excel
    import openpyxl
    if os.path.basename(current_directory) == 'GitHub':
        move_directory = os.path.abspath(os.path.join(current_directory, 'personal_finance'))
        os.chdir(move_directory)
    workbook = openpyxl.load_workbook('$$.xlsx')
    outputs_sheet = workbook['Outputs']

    # Clear K3:N10000
    for row in outputs_sheet.iter_rows(min_row=3, max_row=10000, min_col=11, max_col=14):
        for cell in row:
            cell.value = None

    # Insert upload_df
    start_cell = outputs_sheet.cell(row=3, column=11)
    for row_index, row_data in enumerate(upload_df.values, start=start_cell.row):
        for col_index, cell_value in enumerate(row_data, start=start_cell.column):
            outputs_sheet.cell(row=row_index, column=col_index, value=cell_value)

    workbook.save('$$.xlsx')