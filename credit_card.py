#import setup
from folder_constants import *

#pick the needed df's from folder_constants.py
method = 'credit_card'
original_output_data = original_output_dataframes[method]
upload_df = upload_df_dictionary[method]
original_upload_df = upload_df.copy()

worksheet = spreadsheet.worksheet('Lookup')
data = worksheet.get_all_values()
columns_a_to_c = [row[:3] for row in data]
lookup_df = pd.DataFrame(columns_a_to_c, columns=['Store', 'Category', 'Sub Category'])
lookup_df = lookup_df.drop(0)
lookup_df = lookup_df.sort_values(by=['Category','Store'])
lookup_df = lookup_df.reset_index(drop=True)
lookup_df = lookup_df.drop(0)
duplicates = lookup_df.duplicated(keep='first')
lookup_df = lookup_df[~duplicates]
lookup_df = lookup_df.reset_index(drop=True)
lookup_df['Occurrences'] = lookup_df.groupby('Store')['Store'].transform('count')


#common abbreviations
upload_df['Description'] = upload_df['Description'].apply(lambda x: x.lower() if isinstance(x, str) else x)
upload_df['Description'] = upload_df['Description'].apply(lambda x: 'Amazon' if 'amzn' in x.lower() else x)
upload_df['Description'] = upload_df['Description'].apply(lambda x: 'Amazon' if 'amazon' in x.lower() else x)
upload_df['Description'] = upload_df['Description'].apply(lambda x: 'El Guero Tacos in Tucson' if 'el guero' in x.lower() else x)
upload_df['Description'] = upload_df['Description'].apply(lambda x: 'Tacos Tucson (Street- Taco and Beer Co)' if 'tacos tucson az' in x.lower() else x)
upload_df['Description'] = upload_df['Description'].apply(lambda x: 'The Monica - Tucson' if 'the monica' in x.lower() else x)
upload_df['Description'] = upload_df['Description'].apply(lambda x: 'Instacart' if 'insta wwwbjscom' in x.lower() else x)
upload_df['Description'] = upload_df['Description'].apply(lambda x: 'NFCU Interest' if 'cash adva nces' in x.lower() else x)
upload_df['Description'] = upload_df['Description'].apply(lambda x: 'Pho Chandler' if 'pho chandler chandler az' in x.lower() else x)

#get simplified store
def partial_string_match(s1, s2):
    return s1.lower() in s2.lower()
result_df = upload_df.copy()
for column in lookup_df.columns:
    result_df[column] = upload_df['Description'].apply(lambda x: lookup_df['Store'][lookup_df['Store'].apply(lambda y: partial_string_match(y, x))].iloc[0] if any(lookup_df['Store'].apply(lambda y: partial_string_match(y, x))) else None)
result_df.drop(columns=['Category','Sub Category','Occurrences'],inplace=True) #Bring in all cols, just drop these since they're all the same
result_df['Store'].fillna("", inplace=True)

#check for rows that didn't pull value
nonstore_df = result_df[result_df['Store'] == '']
result_df = result_df[result_df['Store'] != '']
nonstore_df.reset_index(drop=True, inplace=True)
result_df.reset_index(drop=True, inplace=True)

#remove 'chandler','gilbert','tucson','tempe','phoenix','az','arizona' from descriptions
exclusions = ['chandler','gilbert','tucson','tempe','phoenix','az','arizona']
def remove_words(description):
    for word in exclusions:
        description = description.replace(word, '')
    return description
nonstore_df['Description'] = nonstore_df['Description'].apply(remove_words)

#partial match
def find_partial_match(description):
    for part in description.split():
        partial_match = lookup_df.loc[lookup_df['Store'].astype(str).str.contains(part, case=False), 'Store']
        if not partial_match.empty:
            return partial_match.iloc[0]  # Return the first match found
nonstore_df['Partial Match'] = nonstore_df['Description'].apply(find_partial_match)

#fix layout
nonstore_df['Store'] = nonstore_df['Partial Match']
nonstore_df = nonstore_df.drop(columns=['Partial Match'])

#combine dfs again for final result_df
result_df = pd.concat([result_df, nonstore_df], axis=0, ignore_index=True)


#left join to get Category and Sub Category
if len(result_df) > 0:
    result_df = pd.merge(result_df, lookup_df, on='Store', how='left')
    result_df['Category'].fillna("", inplace=True)
    result_df['Sub Category'].fillna("", inplace=True)
    duplicates = result_df.duplicated(keep='first')
    result_df = result_df[~duplicates]
    result_df = result_df.reset_index(drop=True)
    result_df['Transaction Date'] = pd.to_datetime(result_df['Transaction Date'])
    result_df['Spent'] = pd.to_numeric(result_df['Spent'], errors='coerce')
    result_df['Occurrences'] = pd.to_numeric(result_df['Occurrences'], errors='coerce')

    #create a df to filter out duplicate rows (ex: Frys gets duplicated b/c it's Food category and Gas)
    dups_df = result_df[(result_df['Occurrences'] != '') & (result_df['Occurrences'] != 1.0)]
    condition = ((dups_df['Category'] == 'Gas') & (dups_df['Spent'].between(40, 65))) | \
                ((dups_df['Category'] == 'Food') & ((dups_df['Spent'] < 40) | (dups_df['Spent'] > 65)))
    dups_fixed_df = dups_df[condition]

    #delete rows 
    result_df['Occurrences'].fillna("", inplace=True)
    result_df = result_df[(result_df['Occurrences'] == '') | (result_df['Occurrences'] == 1.0)]
    result_df = pd.concat([result_df, dups_fixed_df], ignore_index=True)
    result_df = result_df.sort_values(by=['Transaction Date','Store'], ascending=[False, True])
    result_df = result_df.reset_index(drop = True)
    result_df = result_df.drop(columns='Occurrences')

    # format, add combo col
    original_output_data, upload_df = format_and_combo(original_output_data, result_df)

    # remove rows already stored in output sheet so they're not uploaded twice
    upload_df = remove_rows_already_saved(original_output_data,upload_df)

    #fix amazon categories, by default they are shopping but put these in another df and create a loop that updates each category value, then add back to df
    amazon_df = upload_df[upload_df['Store'] == 'Amazon'].reset_index(drop=True)
    upload_df = upload_df[upload_df['Store'] != 'Amazon'].reset_index(drop=True)

    amazon_df['Sub Category'] = amazon_df['Sub Category'].replace('Amazon','')

    distinct_categories = lookup_df['Category'][lookup_df['Category'] != ''].drop_duplicates().tolist()

    for index, row in amazon_df.iterrows():
        date = pd.to_datetime(row['Transaction Date']).strftime('%m/%d/%y')
        description = row['Description']
        spent = row['Spent']
        store = row['Store']
        distinct_categories = lookup_df['Category'][lookup_df['Category'] != ''].drop_duplicates().tolist()

        print(f"\nOn {date}, there was a ${float(spent):.2f} charge on {store}. What kind of purchase was this?")
        for i, category in enumerate(distinct_categories):
            print(f"{i+1}. {category}")
        selected_index = input("Enter the number corresponding \nto your choice: ")

        selected_category = distinct_categories[int(selected_index) - 1]
        amazon_df.at[index, 'Sub Category'] = selected_category

    #add notes column
    amazon_df['Notes'] = ''
    for index, row in amazon_df.iterrows():
        date = pd.to_datetime(row['Transaction Date']).strftime('%m/%d/%y')
        description = row['Description']
        spent = row['Spent']
        store = row['Store']
        
        user_input = input(f"On {date}, there was a ${float(spent):.2f} charge on {store}. What was this? If no note needed, just press Enter with no text")
        amazon_df.at[index,'Notes'] = user_input


    # add notes column to main df
    non_multistore_df = upload_df[['Transaction Date','Store','Spent','Refunded','Category','Sub Category']]
    non_multistore_df['Notes']=''
    upload_df = pd.concat([non_multistore_df, amazon_df], ignore_index=True)
    upload_df = upload_df.drop(columns='Description')
    upload_df = upload_df.sort_values(by=['Transaction Date','Store'], ascending=[False, True]).reset_index(drop=True)
    #figure out best way to add notes

    # TRAVEL LOGIC
    worksheet = spreadsheet.worksheet('Travel Dates')
    data = worksheet.get_all_values()
    travel_df = pd.DataFrame(data[1:], columns=data[0])
    new_rows = []

    # Iterate through each row in the travel_df
    for index, row in travel_df.iterrows():
        start_date = pd.to_datetime(row['Start Date'], format='%a, %m/%d/%y')
        end_date = pd.to_datetime(row['End Date'], format='%a, %m/%d/%y')
        notes = row['Notes (required)']
        date_range = pd.date_range(start=start_date, end=end_date)
        for date in date_range:
            new_rows.append({'Date': date.strftime('%a, %m/%d/%y'), 'Notes (required)': notes})
    travel_df = pd.DataFrame(new_rows)
    travel_df['Date'] = pd.to_datetime(travel_df['Date'], format='%a, %m/%d/%y')
    travel_df = travel_df.sort_values(by=['Date']).reset_index(drop=True)


    #list of categories that can fall under travel
    travel_categories = ['Drinks','Event','Food','Gas','Self Care','Splurging','Travel','Hotel']

    # add to this list where you know subcategories should not be included in travel expenses
    not_travel_subcategories = ['Home Improvement','Etsy','Pet','Science Center','Games','Car Insurance']


    #look at main_df, if category is in list and transaction date is in travel_df, then change to Travel category and move Category to Sub Category
    upload_df['Transaction Date'] = pd.to_datetime(upload_df['Transaction Date'])
    travel_df['Date'] = pd.to_datetime(travel_df['Date'])

    upload_df = upload_df.merge(travel_df, left_on='Transaction Date', right_on='Date', how='left')
    upload_df = upload_df.drop(columns=['Date'])
    upload_df = upload_df.rename(columns={'Notes (required)':'Travel'})
    upload_df['Travel'].fillna('', inplace=True)

    #if Travel Day = 'Yes' and Category in travel_categories, then make adjustments
    condition = upload_df['Travel'] != ''
    sub_category_present = upload_df['Sub Category'] != ''
    not_in_not_travel_subcategories = upload_df['Sub Category'].isin(not_travel_subcategories)

    upload_df.loc[condition & sub_category_present & ~not_in_not_travel_subcategories, 'Sub Category'] = upload_df['Category'] + ' - ' + upload_df['Sub Category']
    upload_df.loc[condition & ~sub_category_present & ~not_in_not_travel_subcategories, 'Sub Category'] = upload_df['Category']
    upload_df.loc[condition & ~not_in_not_travel_subcategories, 'Category'] = 'Travel'

    #if non travel purchase made during travel, then remove value under Travel column
    mask = (upload_df['Travel'].notnull()) & (upload_df['Category'] != 'Travel')
    upload_df.loc[mask, 'Travel'] = ''


    # get ready for Google Sheets upload
    #if we have a blank store, then pull in the description. Left join the transaction date, spent, and refunded columns to original df
    blank_store = upload_df[upload_df['Store'] == ''].copy().reset_index(drop=True)
    upload_df = upload_df[upload_df['Store'] != ''].reset_index(drop=True)

    # a bunch of datatype casting for a left join..silliness.
    original_upload_df['Transaction Date'] = pd.to_datetime(original_upload_df['Transaction Date'])
    blank_store['Transaction Date'] = pd.to_datetime(blank_store['Transaction Date'])
    original_upload_df['Spent'] = pd.to_numeric(original_upload_df['Spent'], errors='coerce').fillna(0)
    original_upload_df['Refunded'] = pd.to_numeric(original_upload_df['Refunded'], errors='coerce').fillna(0)
    blank_store['Spent'] = pd.to_numeric(blank_store['Spent'], errors='coerce').fillna(0)
    blank_store['Refunded'] = pd.to_numeric(blank_store['Refunded'], errors='coerce').fillna(0)
    blank_store = pd.merge(blank_store, original_upload_df, on=['Transaction Date', 'Spent', 'Refunded'], how='left')
    blank_store['Store'] = blank_store['Description']
    blank_store = blank_store.drop(columns=['Description'])

    #combined df's back together and reorder and reset index
    upload_df = pd.concat([upload_df, blank_store], ignore_index=True)
    upload_df['Transaction Date'] = pd.to_datetime(upload_df['Transaction Date'])
    upload_df = upload_df.sort_values(by=['Transaction Date']).reset_index(drop=True)
    upload_df = upload_df.fillna('')

    #add column that gets category totals partitioned by cycle date
    upload_df['Spent'] = pd.to_numeric(upload_df['Spent'], errors='coerce')
    upload_df['Transaction Date'] = pd.to_datetime(upload_df['Transaction Date'])
    upload_df = upload_df.sort_values(by='Transaction Date')
    

    # we want to combine these new rows with what we originally had pulled from "Credit Card Output", then replace everything in google sheets
    #combine original_output_data with main_df
    if len(upload_df)>0:
        upload_df = pd.concat([upload_df,original_output_data],ignore_index=True)
        if 'combo' in upload_df.columns:
            upload_df = upload_df.drop('combo', axis=1)
        if 'Cycle Totals' in upload_df.columns:
            upload_df = upload_df.drop('Cycle Totals', axis=1)
                                   
        upload_df['Refunded'] = upload_df['Refunded'].replace(0, '')
        
        #rolling 12 months col for pivot table
        upload_df['Transaction Date'] = pd.to_datetime(upload_df['Transaction Date'])
        upload_df['MonthTrunc'] = upload_df['Transaction Date'].dt.to_period('M')
        max_month = upload_df['MonthTrunc'].max()
        upload_df['Rolling12'] = ['Yes' if (max_month - x).n <= 12 else 'No' for x in upload_df['MonthTrunc']]
        #drop monthtrunc
        upload_df = upload_df.drop(columns='MonthTrunc')

        #format and sort
        upload_df = upload_df.sort_values(by=['Transaction Date','Travel']).reset_index(drop=True)
        upload_df['Transaction Date'] = upload_df['Transaction Date'].astype(str)
        upload_df['Spent'] = upload_df['Spent'].astype(float)

        #upload
        worksheet = spreadsheet.worksheet('Outputs')
        worksheet.range(sheets_info_dict[method][1]).clear()
        data = upload_df.values.tolist()
        worksheet.update(sheets_info_dict[method][0], data)

        
        ### Add pivot data
        #get totals by category and date, for pivot tables
        pivot_df = upload_df.copy()
        # pivot_df = pivot_df.drop(columns=['Store','Refunded','Notes','Travel','Cycle Totals','Rolling12'])
        pivot_df['Transaction Date'] = pd.to_datetime(pivot_df['Transaction Date'])
        pivot_df['Spent'] = pd.to_numeric(pivot_df['Spent'], errors='coerce').fillna(0)
        pivot_df['Average Spent by Day'] = pivot_df.groupby(['Transaction Date', 'Category'])['Spent'].transform('mean').round(2)
        pivot_df = pivot_df.drop(columns=['Spent'])
        pivot_df = pivot_df.groupby(['Transaction Date', 'Category'])['Average Spent by Day'].sum().reset_index()
        pivot_df = pivot_df.sort_values(by='Transaction Date')

        pivot_df['Cycle Start'] = pivot_df['Transaction Date'].dt.to_period('M')
        pivot_df['Cycle Total'] = pivot_df.groupby(['Cycle Start', 'Category'])['Average Spent by Day'].cumsum()
        pivot_df = pivot_df.drop(columns='Cycle Start')


        first_day_prior_month = pd.Timestamp.now() - pd.DateOffset(months=1)
        first_day_prior_month = first_day_prior_month.replace(day=1)
        date_range = pd.date_range(first_day_prior_month, pd.Timestamp.now() - pd.DateOffset(1), freq='D').date
        category_subcategory_combinations = pivot_df[['Category']].drop_duplicates()
        all_combinations = pd.MultiIndex.from_product([date_range, category_subcategory_combinations['Category']], names=['Transaction Date', 'Category']).to_frame(index=False).reset_index(drop=True)
        all_combinations['Transaction Date'] = pd.to_datetime(all_combinations['Transaction Date'])
        final_pivot = pd.merge(all_combinations, pivot_df, how='left', on=['Transaction Date', 'Category'])
        final_pivot = final_pivot.drop_duplicates()
        final_pivot = final_pivot.drop(columns='Average Spent by Day')
        final_pivot['Cycle Total'] = final_pivot.groupby(['Category'])['Cycle Total'].fillna(method='ffill')
        final_pivot = final_pivot.dropna(subset=['Cycle Total']).reset_index(drop=True)

        #pull in budget
        final_pivot = pd.merge(final_pivot, budget_df, left_on='Category', right_on='Category', how='left')
        final_pivot['Days in Month'] = final_pivot['Transaction Date'].dt.days_in_month
        final_pivot['Day Number'] = final_pivot['Transaction Date'].dt.day
        final_pivot['Budget'] = final_pivot.apply(lambda row: row['Budget'] / row['Days in Month'] * row['Day Number'] if row['Type'] == 'Gradual' else row['Budget'], axis=1).round(2).fillna(0)
        final_pivot = final_pivot.drop(columns=['Type','Days in Month','Day Number'])
        final_pivot = final_pivot.rename(columns={'Cycle Total':'Spent'})


        # Update excel
        import openpyxl
        if os.path.basename(current_directory) == 'GitHub':
            move_directory = os.path.abspath(os.path.join(current_directory, 'personal_finance'))
            os.chdir(move_directory)

        workbook = openpyxl.load_workbook('$$.xlsx')
        outputs_sheet = workbook['Outputs']

        # Clear A3:E10000
        for row in outputs_sheet.iter_rows(min_row=3, max_row=10000, min_col=1, max_col=upload_df.shape[1]):
            for cell in row:
                cell.value = None

        # Insert upload_df
        start_cell = outputs_sheet.cell(row=3, column=1)
        for row_index, row_data in enumerate(upload_df.values, start=start_cell.row):
            for col_index, cell_value in enumerate(row_data, start=start_cell.column):
                outputs_sheet.cell(row=row_index, column=col_index, value=cell_value)
        
        #Pivot page upload
        outputs_sheet = workbook['Pivot Data']
        # Clear A3:E10000
        for row in outputs_sheet.iter_rows(min_row=1, max_row=10000, min_col=1, max_col=final_pivot.shape[1]):
            for cell in row:
                cell.value = None

        # Insert upload_df
        start_cell = outputs_sheet.cell(row=2, column=1)
        for row_index, row_data in enumerate(final_pivot.values, start=start_cell.row):
            for col_index, cell_value in enumerate(row_data, start=start_cell.column):
                if isinstance(cell_value, pd.Timestamp):
                    cell_value = cell_value.strftime('%Y-%m-%d')
                outputs_sheet.cell(row=row_index, column=col_index, value=cell_value)


        workbook.save('$$.xlsx')
